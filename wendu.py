#coding=gbk
import os
import re
import openpyxl
import sys
import time
from openpyxl.styles import Alignment 
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Side, Border
import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
import tkinter.font as tkFont
import shutil
import pyexcel as p



ORGIN_FILE_PATH = ""
BGN_ANALYZ_TIME = ""
END_ANALYZ_TIME = ""
GEN_FILE_PATH = ""

master = tk.Tk()








#!=================================================================
#! Function: TemperatureAnalyz
#!================================================================= 
def TemperatureAnalyze():
    global ORGIN_FILE_PATH
    global BGN_ANALYZ_TIME
    global END_ANALYZ_TIME
    global GEN_FILE_PATH

    master.title("萌：室温平均值计算")
    sw = master.winfo_screenwidth()
    sh = master.winfo_screenheight()
    ww = 320
    hh = 250
    x  = (sw - ww)/2
    y  = (sh - hh)/2
    master.geometry('%dx%d+%d+%d' %(ww,hh,x,y))

    #master.geometry("500x600")

    tk.Label(master, text="原始表格路径").grid(row=0,column=0, sticky="E", padx=10, pady=10)
    tk.Label(master, text="开始测量时间").grid(row=1,column=0, sticky="E", padx=10, pady=10)
    tk.Label(master, text="结束测量时间").grid(row=2,column=0, sticky="E", padx=10, pady=10)
    tk.Label(master, text="表格生成路径").grid(row=3,column=0, sticky="E", padx=10, pady=10)
    #fontStyle = tkFont.Font(family="Lucida Grande", size=10)
    #label3 = tk.Label(master, text="表格生成路径", font=fontStyle)
    #label3.grid(row=3,column=0, sticky="E", padx=10, pady=10)
    
    entry_txt1 = tk.StringVar(value="")
    e1 = tk.Entry(master, textvariable=entry_txt1)
    e2 = tk.Entry(master)
    e3 = tk.Entry(master)
    entry_txt4 = tk.StringVar(value="") 
    e4 = tk.Entry(master, textvariable=entry_txt4)

    e1.grid(row=0,column=1,padx=10, pady=10)
    e2.grid(row=1,column=1,padx=10, pady=10)
    e3.grid(row=2,column=1,padx=10, pady=10)
    e4.grid(row=3,column=1,padx=10, pady=10)


    #--------------------------------------------------
    def tanchuang():
        
        ORGIN_FILE_PATH = e1.get()
        BGN_ANALYZ_TIME = e2.get()  
        END_ANALYZ_TIME = e3.get()
        GEN_FILE_PATH   = e4.get()

        confirm_message = \
            "原始表格路径: " +ORGIN_FILE_PATH + "\n" + \
            "开始测量时间: " +BGN_ANALYZ_TIME + "\n" + \
            "结束测量时间: " +END_ANALYZ_TIME + "\n" + \
            "表格生成路径: " +GEN_FILE_PATH


        rt_ask_msg = messagebox.askokcancel(title="信息确认", message=confirm_message)
        if rt_ask_msg:
            DealOrignExcelFile(ORGIN_FILE_PATH, BGN_ANALYZ_TIME, END_ANALYZ_TIME, GEN_FILE_PATH) 
            master.quit()
    
    def setEntry1Txt():
        askpath = filedialog.askdirectory()
        entry_txt1.set(askpath)

    def setEntry4Txt():
        askpath = filedialog.askdirectory()
        entry_txt4.set(askpath)
        
    
    driver = tk.Button(master,text="开始计算",width=10, command=tanchuang)
    #driver.grid(row=4, column=2, sticky="sw", padx=10, pady=5)
    driver.place(relx=0.5, rely=0.9, anchor="center")

    choose_path_button = tk.Button(master, text="路径",width=3,height=1,justify="left",command=setEntry1Txt)
    choose_path_button.grid(row=0,column=2,padx=5,pady=5)

    choose_path_button = tk.Button(master, text="路径",width=3,height=1,justify="left",command=setEntry4Txt)
    choose_path_button.grid(row=3,column=2,padx=5,pady=5)

    master.mainloop()























#!=================================================================
#! Function: Deal Original Excel File 
#!================================================================= 
def get_choose_data(excel_file, start_time, end_time):
    new_change = 0

    if re.search("\.xls$", excel_file):
        new_change = 1
        #shutil.copy(excel_file, excel_file + "x")
        p.save_book_as(file_name=excel_file, dest_file_name=(excel_file + "x"))
        excel_file = excel_file + "x"
        data_sheet_name = "report"
        time_column = 3
        data_column = 4
    else:
        data_sheet_name = "Sheet1"
        time_column = 1
        data_column = 2

    wb = openpyxl.load_workbook(excel_file)
    sheet = wb[data_sheet_name]
    #sys.exit("-------------")
    
    
    data_list = []
    match_flag = 0
    for i in range(1, sheet.max_row):
        #print (sheet.cell(i,time_column).value)
        #print (start_time, str(sheet.cell(i,1).value).strip())
        #break
        if start_time in str(sheet.cell(i,time_column).value): # FIXME : Format is 2022-02-23 01:30:00
            #print ("match")
            match_flag = 1
        if match_flag:
            test_value = float(sheet.cell(i,data_column).value)
            #print (sheet.cell(i,1).value, sheet.cell(i,2).value)
            data_list.append(test_value)
        if end_time in str(sheet.cell(i,time_column).value):
            break
    
    if new_change == 1:
        os.remove(excel_file)

    return data_list


def write_sheet_cell(Sheet, Row, Column, Write_Value):
    cur_cell = Sheet.cell(Row, Column, Write_Value)
    cur_cell.alignment = Alignment(horizontal='center', vertical='center')
    cur_cell.number_format = '0.0'
    
    border_cofig = Side(border_style='thin',color=colors.BLACK)
    border = Border(left=border_cofig, right=border_cofig, top=border_cofig, bottom=border_cofig)
    cur_cell.border = border

    return cur_cell



def get_time_lst(excel_file, start_time, end_time):
    TIME_LST = []
    if len(TIME_LST) > 0:
        return
    
    new_change = 0
    if re.search("\.xls$", excel_file):
        new_change = 1
        #shutil.copy(excel_file, excel_file + "x")
        p.save_book_as(file_name=excel_file, dest_file_name=(excel_file + "x"))
        excel_file = excel_file + "x"
        data_sheet_name = "report"
        time_column = 3
        data_column = 4
    else:
        data_sheet_name = "Sheet1"
        time_column = 1
        data_column = 2


    wb = openpyxl.load_workbook(excel_file)
    sheet = wb[data_sheet_name]

    match_flag = 0
    for i in range(1, sheet.max_row):
        if start_time in str(sheet.cell(i,time_column).value):
            TIME_LST = []
            match_flag = 1
        if match_flag:
            TIME_LST.append(str(sheet.cell(i,time_column).value).split()[-1])
        if end_time in str(sheet.cell(i,time_column).value):
            #print (TIME_LST)
            break

    if new_change == 1:
        os.remove(excel_file)

    return TIME_LST





def DealOrignExcelFile(orgin_file_path, bgn_analyz_time, end_analyz_time, gen_file_path):

    #print( orgin_file_path)
    #print( bgn_analyz_time)
    #print( end_analyz_time)
    #print( gen_file_path  )

    excel_file  = orgin_file_path
    start_time  = bgn_analyz_time
    end_time    = end_analyz_time
    GEN_FILE    = gen_file_path + "/" + time.strftime("%y%m%d%H%M_summary.xlsx", time.localtime())


    


    TIME_LST = []
    #print (GEN_FILE)

    #print (os.listdir(excel_file))

    if not os.path.isdir(excel_file):
        info = ("%0s 不存在!" % excel_file)
        messagebox.showwarning(title="文件夹不存在", message=info)
        TemperatureAnalyze()
        return
    
    #wait_window = tk.Tk()
    #tk.Label(wait_window, text="开始处理表格了，请稍等").grid(row=0,column=0, sticky="E", padx=10, pady=10)


    dirs_list = []
    time_list = []
    data_dict = {}
    for d in os.listdir(excel_file):
        dir_name = excel_file + "\\" + d
        if not os.path.isdir(dir_name):
            continue
        
        print ("\n\n\n======================================================================")
        print ("进入文件夹：%0s" % dir_name)
        dirs_list.append(d)
        array_data_list = []
        for xlsx in os.listdir(dir_name):
            if not re.search("\.(xls)|(xlsx)$", xlsx):
                continue

            print ("读取表格：%0s" % xlsx)
            xlsx_name = dir_name + "\\" + xlsx
            xlsx_data_dict = {}
            single_data_lst = []
            if os.path.isfile(xlsx_name):
                #print (xlsx_name)
                single_data_lst = get_choose_data(xlsx_name, start_time, end_time)
                if len(single_data_lst) == 0:
                    warn_info = "文件 %0s 没有 %0s ~ %0s 时间段的数据" % (xlsx_name, start_time, end_time)
                    messagebox.showwarning(title="数据缺失", message=warn_info)
                    continue
                    
                xlsx_data_dict[xlsx] = get_choose_data(xlsx_name, start_time, end_time)
                TIME_LST = get_time_lst(xlsx_name, start_time, end_time)
                #print (xlsx_name, xlsx_data_dict)
                array_data_list.append(xlsx_data_dict)
                #print ()
        #print (len(array_data_list))
        data_dict[d] = array_data_list;
        #print (data_dict)

    #print ("Dict size =%0d" % len(data_dict))



    new_wb = openpyxl.Workbook()
    #new_sheet = new_wb.create_sheet("Sheet1")
    new_sheet = new_wb[new_wb.sheetnames[0]]
    column = 1


    # write "时间"
    row = 2
    tmstr =  "date"
    write_sheet_cell(new_sheet, row, column, tmstr)
    
    # Write time in xlsx
    for t in TIME_LST:
        row += 1
        write_sheet_cell(new_sheet, row, column, t)
    row += 1
    cur_cell = write_sheet_cell(new_sheet, row, column, "average")
    cur_cell.fill = PatternFill("solid", fgColor=colors.Color(indexed=40))
    column += 1
    row = 1

    #cell_color_lst = [colors.RED, colors.GREEN, colors.BLUE]
    color_rnd = 40

    for key in data_dict.keys():
        #print ("-----------------------------------------------------")
        #print (key, len(data_dict[key]))
        dir_size = len(data_dict[key])
        color_rnd += 1

        new_sheet.merge_cells(start_row=1, start_column=column, end_row=1, end_column=(column + dir_size - 1))
        cur_cell = write_sheet_cell(new_sheet, 1, column, key)
        cur_cell.fill = PatternFill("solid", fgColor=colors.Color(indexed=color_rnd))
        row = 2
        for lst_dict in data_dict[key]:
            x = list(lst_dict.keys())[0]
            #print (x, lst_dict)

            cur_cell = write_sheet_cell(new_sheet, row, column, x.replace(".xlsx", ""))
            cur_cell.fill = PatternFill("solid", fgColor=colors.Color(indexed=color_rnd))
            for dt in lst_dict[x]:
                row += 1
                write_sheet_cell(new_sheet, row, column, dt)

            #! write average value at last
            row += 1
            cur_cell = write_sheet_cell(new_sheet, row, column, sum(lst_dict[x])/float(len(lst_dict[x])))
            cur_cell.fill = PatternFill("solid", fgColor=colors.Color(indexed=40))

            column += 1 
            row = 2
    new_wb.save(GEN_FILE)

    #wait_window.destroy()
    
    info = "\n\n分析表格已经生成：\n%0s" % GEN_FILE
    messagebox.showinfo(title="数据缺失", message=info)
    
            
    #if os.path.isfile(excel_file):
    #    print (excel_file)






#!=================================================================
#! Function: build_phase
#!================================================================= 
if __name__ == "__main__":

    TemperatureAnalyze()


